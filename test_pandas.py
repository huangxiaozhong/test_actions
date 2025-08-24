from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------
# 1. 创建一个新的 Excel 工作簿
# --------------------------
wb = Workbook()
# 获取默认的工作表（第一个工作表）
ws = wb.active
# 重命名工作表（可选）
ws.title = "学生成绩表"

# --------------------------
# 2. 写入表头（第一行）
# --------------------------
headers = ["姓名", "年龄", "语文成绩", "数学成绩", "总分"]
ws.append(headers)  # 按行追加数据（表头）

# --------------------------
# 3. 写入学生数据（多行）
# --------------------------
students = [
    ["张三", 18, 90, 85, 175],
    ["李四", 17, 88, 92, 180],
    ["王五", 19, 95, 89, 184],
    ["赵六", 18, 82, 78, 160]
]
for student in students:
    ws.append(student)  # 按行追加数据

# --------------------------
# 4. 设置格式（可选，美化表格）
# --------------------------
# 设置表头字体（加粗、蓝色）
header_font = Font(bold=True, color="0000FF")
for col in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")  # 居中对齐

# 设置总分列（第5列）的背景色（浅灰色）
fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
for row in range(2, len(students)+2):  # 数据从第2行开始（表头是第1行）
    cell = ws.cell(row=row, column=5)
    cell.fill = fill

# 调整列宽（根据内容自动适应，或手动设置）
for col in ws.columns:
    max_length = 0
    column = get_column_letter(col[0].column)  # 获取列字母（如 "A"）
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1.2  # 调整宽度（经验值）
    ws.column_dimensions[column].width = adjusted_width

# --------------------------
# 5. 保存 Excel 文件
# --------------------------
wb.save("学生成绩.xlsx")
print("Excel 文件已生成：学生成绩.xlsx")
